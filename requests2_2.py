import os
import re
import time
import traceback
from datetime import datetime

import pandas as pd
import pytz
from dotenv import load_dotenv
from lxml import html
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

# # Load environment variables from a .env file
# load_dotenv()

# # Now you can use os.getenv to read the environment variables
# n_id = os.getenv("NAVER_ID")
# n_pw = os.getenv("NAVER_PW")

# make driver
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

# Website URL
url = "https://nid.naver.com/nidlogin.login"
posts_url = "https://cafe.naver.com/noljatravel?iframe_url=/ArticleList.nhn%3Fsearch.clubid=26681849%26userDisplay=50%26search.boardtype=L%26search.specialmenutype=%26search.totalCount=501%26search.cafeId=26681849%26search.page="
# naver_url = "https://cafe.naver.com"

# # Access the website
# driver.get(url)
# time.sleep(5)  # You can adjust the sleep time as needed

# login_input = driver.find_element(By.ID, "id")
# login_input.clear()
# login_input.send_keys(n_id)
# time.sleep(1)  # You can adjust the sleep time as needed
# pw_input = driver.find_element(By.ID, "pw")
# pw_input.clear()
# pw_input.send_keys(n_pw)
# time.sleep(1)  # You can adjust the sleep time as needed
# # Assuming 'driver' is your WebDriver instance

# button = driver.find_element(By.ID, "log.login")
# button.click()
# time.sleep(30)  # Time to login
page = 0
status = True
article_list = []

pattern_S = ["모델 S", "모델 s", "모델S", "모델s"]
pattern_S = "|".join(pattern_S)
S_article_list = []

pattern_X = ["모델 X", "모델 x", "모델X", "모델x"]
pattern_X = "|".join(pattern_X)
X_article_list = []

pattern_Y = ["모델 Y", "모델 y", "모델Y", "모델y"]
pattern_Y = "|".join(pattern_Y)
Y_article_list = []

pattern_3 = ["모델 3", "모델 3", "모델3", "모델3"]
pattern_3 = "|".join(pattern_3)
t_article_list = []

while status:
    page += 1
    driver.get(posts_url + str(page))

    try:
        iframe = driver.find_element(By.ID, "cafe_main")
        driver.switch_to.frame(iframe)
    except:
        pass

    post_table = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[4]")
    tbody = post_table.find_element(By.TAG_NAME, "tbody")
    rows = tbody.find_elements(By.XPATH, "./tr")
    print(f"page: {page}")

    # Now you can iterate over each row
    time.sleep(5)
    for row in rows:
        try:
            article = row.find_element(By.CLASS_NAME, "td_article")
            board_name = article.find_element(By.CLASS_NAME, "board-name").text
            inner_list = article.find_element(By.CLASS_NAME, "inner_list")
            # Find the first <a> tag within 'inner_list'
            first_a_tag = inner_list.find_element(By.TAG_NAME, "a")
            # Get the 'href' attribute of the first <a> tag
            url_r = first_a_tag.get_attribute("href")

            title = inner_list.find_element(By.TAG_NAME, "a").text
            try:
                cmt = inner_list.find_element(By.CSS_SELECTOR, "a.cmt").text
                cmt = int(cmt.replace("[", "").replace("]", ""))
            except NoSuchElementException:
                cmt = 0

            name = row.find_element(By.CLASS_NAME, "p-nick").text
            time_ = row.find_element(By.CLASS_NAME, "td_date").text
            if len(time_) > 6:
                status = False
                break
            view = row.find_element(By.CLASS_NAME, "td_view").text
            view = int(view.replace(",", ""))
            # if view

            temp_list = [time_, board_name, title, view, cmt, name, url]
            article_list.append(temp_list)
            if re.search(pattern_S, title):
                S_article_list.append(temp_list)
            if re.search(pattern_X, title):
                X_article_list.append(temp_list)
            if re.search(pattern_Y, title):
                Y_article_list.append(temp_list)
            if re.search(pattern_3, title):
                t_article_list.append(temp_list)

        except Exception as e:
            print(row)
            print(f"Error: {e}")
            print("Traceback:")
            traceback.print_exc()
            pass

columns = ["시간", "게시판", "제목", "조회수", "댓글수", "게시자", "url"]

df_all = pd.DataFrame(data=article_list, columns=columns)
df_S = pd.DataFrame(data=S_article_list, columns=columns)
df_X = pd.DataFrame(data=X_article_list, columns=columns)
df_Y = pd.DataFrame(data=Y_article_list, columns=columns)
df_3 = pd.DataFrame(data=t_article_list, columns=columns)

# Korean timezone
korean_tz = pytz.timezone("Asia/Seoul")

# Current time in Korean timezone
now_korean = datetime.now(korean_tz)
date = now_korean.strftime("%Y/%m/%d")
time = now_korean.strftime("%H:%M")

# Specify the Excel file path
file_path = "cafe3.xlsx"


def add_sheet_and_data(writer, sheet_name, df):
    ws = writer.book.create_sheet(title=sheet_name)
    # Write metadata
    ws["A1"] = "Date"
    ws["B1"] = date
    ws["A2"] = "Time"
    ws["B2"] = time
    # Add DataFrame data
    df.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False)

    # Adjust column widths based on the data
    for column_cells in ws.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = (
            max_length + 2
        )  # Adding a little extra space


with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    # Dictionary of sheet names and corresponding DataFrames
    sheets_data = {
        "All": df_all,
        "모델 S": df_S,
        "모델 X": df_X,
        "모델 Y": df_Y,
        "모델 3": df_3,
    }

    # Loop through the dictionary to create sheets and add data
    for sheet_name, df in sheets_data.items():
        add_sheet_and_data(writer, sheet_name, df)

# No need to call writer.save() inside the context manager block

print(f"File '{file_path}' has been saved with Korean time metadata.")
